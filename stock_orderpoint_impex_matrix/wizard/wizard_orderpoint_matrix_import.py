# Copyright 2020 Akretion
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

import base64
from io import BytesIO

import openpyxl

from odoo import _, fields, models
from odoo.exceptions import ValidationError

from . import common as CONSTANTS


class WizardOrderpointMatrixImport(models.TransientModel):
    _name = "wizard.orderpoint.matrix.import"
    _description = "Wizard Orderpoint Matrix Import"

    filename = fields.Char()
    file = fields.Binary()

    def _find_real_last_column(self, worksheet):
        """
        The last column and row are actually written in the excel file
        Openpyxl doesn't automatically validate if it is right or not
        """
        tentative_last_column = worksheet.max_column
        for col in reversed(range(tentative_last_column)):
            if worksheet.cell(row=2, column=col + 1).value:
                break
        return col + 1

    def _find_real_last_row(self, worksheet, max_col):
        """ See _find_real_last_column """
        tentative_last_row = worksheet.max_row
        for row in reversed(range(tentative_last_row)):
            row_has_val = any(
                worksheet.cell(row + 1, col + 1).value for col in range(max_col)
            )
            if row_has_val:
                break
        return row + 1

    def _validate_by_warehouse_names(self, sheet, number_of_wh):
        positions = [
            CONSTANTS.COLUMN_START_WH_BLOCKS + itr * CONSTANTS.LEN_COLUMNS_PER_WH
            for itr in range(number_of_wh)
        ]
        wh_names = []
        for position in positions:
            prefixed_name = sheet.cell(row=1, column=position).value
            name = prefixed_name.split(CONSTANTS.PREFIX_HEADER_WH)[1]
            wh_names.append(name)
        warehouses = self.env["stock.warehouse"]
        for name in wh_names:  # note by doing it 1 by 1 we preserve list order
            candidate = self.env["stock.warehouse"].search([("name", "=", name)])
            if len(candidate.ids) != 1:
                raise ValidationError(
                    _("Warehouse names should match to exactly one warehouse")
                )
            warehouses += candidate
        return warehouses

    def _validate_by_column_parity(self, last_column):
        if (
            last_column - (CONSTANTS.COLUMN_START_WH_BLOCKS - 1)
        ) % CONSTANTS.LEN_COLUMNS_PER_WH != 0:
            raise ValidationError(
                _("Bad parity for columns, some were removed or added")
            )

    def _validate_warehouses(self, sheet):
        last_column = self._find_real_last_column(sheet)
        self._validate_by_column_parity(last_column)
        number_of_wh = (last_column - 1) // CONSTANTS.LEN_COLUMNS_PER_WH
        return self._validate_by_warehouse_names(sheet, number_of_wh)

    def _validate_products(self, sheet):
        last_row = self._find_real_last_row(sheet, 1)
        product_codes = [
            sheet.cell(column=1, row=row).value
            for row in range(CONSTANTS.ROW_START_PRODUCTS, last_row + 1)
        ]
        products = self.env["product.product"]
        for (
            product_code
        ) in product_codes:  # note by doing it 1 by 1 we preserve list order
            candidate = self.env["product.product"].search(
                [("default_code", "=", product_code)], limit=1
            )
            if len(candidate.ids) != 1:
                raise ValidationError(
                    _("Product codes should match to exactly one product")
                )
            products += candidate
        return products

    def _validate_excel(self, sheet):
        warehouses = self._validate_warehouses(sheet)
        products = self._validate_products(sheet)
        return warehouses, products

    def _build_orderpoint_matrix(self, sheet, warehouses, products):
        """
        returns a matrix of format:
        [
            [[row1block1], [row1block2], ...]
            [[row2block1], [row2block2], ...]
        ]
        Where each block is a list of vals from corresponding to MAPPINGS_COLUMNS_PER_WH
        """
        result = []
        no_of_blocks = len(warehouses.ids)
        for row, _product in enumerate(products, start=CONSTANTS.ROW_START_PRODUCTS):
            row_vals = []
            for idx_block in range(no_of_blocks):
                col_block_start = (
                    idx_block * CONSTANTS.LEN_COLUMNS_PER_WH
                    + CONSTANTS.COLUMN_START_WH_BLOCKS
                )
                block_vals = [
                    sheet.cell(row=row, column=col_block_start + col_itr).value
                    for col_itr in range(CONSTANTS.LEN_COLUMNS_PER_WH)
                ]
                have_values = []
                for el in block_vals[1:]:
                    have_values.append(el is None or el == "")
                if any(have_values) and not all(have_values):
                    raise ValidationError(
                        _(
                            "For each warehouse, either fill all values or empty all values"
                        )
                    )
                row_vals.append(block_vals)
            result.append(row_vals)
        return result

    def _match_orderpoint(self, product, warehouse):
        orderpoint = self.env["stock.warehouse.orderpoint"].search(
            [
                ("product_id", "=", product.id),
                ("warehouse_id", "=", warehouse.id),
                ("location_id", "=", warehouse.lot_stock_id.id),
            ]
        )
        return orderpoint

    def _update_or_delete_orderpoint(self, orderpoint, vals):
        empty_line = all([val is None or val == "" for val in vals])
        if empty_line:
            orderpoint.unlink()
        else:
            orderpoint.product_min_qty = vals[1]
            orderpoint.product_max_qty = vals[2]
            orderpoint.lead_days = vals[3]
            orderpoint.qty_multiple = vals[4]

    def _process_sheet(self, sheet, warehouses, products):
        all_rows_vals = self._build_orderpoint_matrix(sheet, warehouses, products)
        for idx_row, row_vals in enumerate(all_rows_vals):
            for idx_block, block_vals in enumerate(row_vals):
                product = products[idx_row]
                warehouse = warehouses[idx_block]
                orderpoint = self._match_orderpoint(product, warehouse)
                if orderpoint:
                    self._update_or_delete_orderpoint(orderpoint, block_vals)
                else:
                    self.env["stock.warehouse.orderpoint"].create(
                        {
                            "name": product.name + " (" + warehouse.name + ")",
                            "product_id": product.id,
                            "warehouse_id": warehouse.id,
                            "location_id": warehouse.lot_stock_id.id,
                            "product_min_qty": block_vals[1],
                            "product_max_qty": block_vals[2],
                            "lead_days": block_vals[3],
                            "qty_multiple": block_vals[4],
                        }
                    )

    def button_import_excel(self):
        wb = openpyxl.load_workbook(
            BytesIO(base64.b64decode(self.file.decode("utf-8")))
        )
        sheet = wb.worksheets[0]
        warehouses, products = self._validate_excel(sheet)
        self._process_sheet(sheet, warehouses, products)
        return True
