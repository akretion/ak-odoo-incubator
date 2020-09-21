# Copyright 2020 Akretion
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

import base64
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Alignment
from odoo import fields, models, _
from . import common as CONSTANTS


class WizardOrderpointMatrixExport(models.TransientModel):
    _name = "wizard.orderpoint.matrix.export"
    _description = "Wizard Orderpoint Matrix Export"

    dummy_export_done = fields.Boolean()
    filename = fields.Char()
    file = fields.Binary()
    warehouse_ids = fields.Many2many("stock.warehouse", required=True)
    product_ids = fields.Many2many("product.product", compute="_compute_product_ids")

    def _compute_product_ids(self):
        locations = self.warehouse_ids.mapped("lot_stock_id")
        orderpoints = self.env["stock.warehouse.orderpoint"].search(
            [("location_id", "in", locations.ids)]
        )
        self.product_ids = orderpoints.mapped("product_id")

    def _generate_orderpoint_vals(self, product, warehouse):
        return ["" for field in CONSTANTS.MAPPING_COLUMNS_PER_WH]

    def _search_orderpoint_vals(self, product, warehouse):
        orderpoint = self.env["stock.warehouse.orderpoint"].search(
            [
                ("product_id", "=", product.id),
                ("warehouse_id", "=", warehouse.id),
                ("location_id", "=", warehouse.lot_stock_id.id),
            ]
        )
        result = False
        if orderpoint:
            result = [
                orderpoint.qty_multiple,
                orderpoint.lead_days,
                orderpoint.product_id.with_context(
                    {"warehouse": warehouse.id}
                ).qty_available,
                orderpoint.product_min_qty,
                orderpoint.product_max_qty,
            ]
        return result

    def _get_orderpoint_vals(self, products, warehouse_ids):
        """
        For each warehouse, searches the matching orderpoint. If it exists,
        populate result with the existing values. Or else, put in empty values
        Format: [
            [product1_name, wh1field1, wh1field2, ... , wh2field1, wh2field2 ...],
            [product2_name, wh1field1, wh1field2, ... , wh2field1, wh2field2 ...]
        ...
        ]
        """
        result = []
        for product in products:
            row = [product.default_code, product.name]
            for warehouse in warehouse_ids:
                row += self.with_context(
                    {"warehouse": warehouse.id}
                )._search_orderpoint_vals(
                    product, warehouse
                ) or self._generate_orderpoint_vals(
                    product, warehouse
                )
            result.append(row)
        return result

    def _write_headers(self, sheet):
        sheet.cell(column=1, row=2).value = _("Code Article")
        sheet.cell(column=2, row=2).value = _("Article")
        for idx_col_wh, warehouse in enumerate(
            self.warehouse_ids.sorted(lambda r: r.id), start=0
        ):
            block_start = (
                CONSTANTS.COLUMN_START_WH_BLOCKS
                + idx_col_wh * CONSTANTS.LEN_COLUMNS_PER_WH
            )
            sheet.cell(column=block_start, row=1).value = (
                CONSTANTS.PREFIX_HEADER_WH + warehouse.name
            )
            for idx_cell, header in enumerate(
                CONSTANTS.MAPPING_COLUMNS_PER_WH.values(), start=0
            ):
                sheet.cell(column=block_start + idx_cell, row=2).value = header

    def _fill_data(self, sheet):
        rows = self._get_orderpoint_vals(
            self.product_ids.sorted(lambda r: r.id),
            self.warehouse_ids.sorted(lambda r: r.id),
        )
        for idx_row, row_data in enumerate(rows, start=3):
            for idx_cell, cell_data in enumerate(row_data, start=1):
                sheet.cell(column=idx_cell, row=idx_row).value = cell_data

    def _apply_formatting_misc(self, sheet):
        # Freeze panes
        sheet.freeze_panes = "A2"
        # Set first columns widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 20
        alternate_grey = False
        for idx_block, warehouse in enumerate(self.warehouse_ids):
            start_of_block = (
                CONSTANTS.COLUMN_START_WH_BLOCKS
                + idx_block * CONSTANTS.LEN_COLUMNS_PER_WH
            )
            end_of_block = start_of_block + CONSTANTS.LEN_COLUMNS_PER_WH
            # Set column width
            for itr in range(CONSTANTS.LEN_COLUMNS_PER_WH):
                letter = openpyxl.utils.cell.get_column_letter(start_of_block + itr)
                sheet.column_dimensions[letter].width = 20
            # Merge warehouse cells
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=start_of_block,
                end_column=end_of_block - 1,
            )
            # Center warehouse cells
            sheet.cell(row=1, column=start_of_block).alignment = Alignment(
                horizontal="center"
            )
            # Alternate colors between warehouse blocks
            alternate_grey = not alternate_grey
            grey_fill = PatternFill(fill_type="lightGray")
            if alternate_grey:
                for idx_col in range(start_of_block, end_of_block):
                    for idx_row in range(1, CONSTANTS.ROW_MAX_STYLING):
                        sheet.cell(row=idx_row, column=idx_col).fill = grey_fill

    def _refresh_xlsx_file(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        self._write_headers(sheet)
        self._fill_data(sheet)
        self._apply_formatting_misc(sheet)
        tmpfile = BytesIO()
        wb.save(tmpfile)
        self.file = base64.b64encode(tmpfile.getvalue())
        return True

    def button_export_refresh_result(self):
        self._refresh_xlsx_file()
        self.dummy_export_done = True
        return {
            "name": _("Orderpoint export"),
            "view_type": "form",
            "view_mode": "form",
            "views": [
                (
                    self.env.ref(
                        "stock_orderpoint_impex_matrix.wizard_orderpoint_matrix_export_view_form"
                    ).id,
                    "form",
                )
            ],
            "res_model": "wizard.orderpoint.matrix.export",
            "type": "ir.actions.act_window",
            "res_id": self.id,
            "target": "new",
        }
